import os
from typing import Union, TextIO, Optional, List

import click

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

from linkml.utils.generator import Generator, shared_arguments
from linkml_runtime.linkml_model.meta import (
    SchemaDefinition,
    ClassDefinition,
    EnumDefinition,
    PermissibleValue,
    PermissibleValueText,
    SlotDefinition,
)
from linkml_runtime.utils.formatutils import camelcase


class ExcelGenerator(Generator):
    """This class is a blueprint for the generator module that is responsible
    for automatically creating Excel spreadsheets from the LinkML schema.

    :param schema: LinkML schema object
    :type schema: class:`SchemaDefinition`
    :param output: LinkML schema specification in YAML format
    :type output: str
    """

    generator_name = os.path.splitext(os.path.basename(__file__))[0]
    generator_version = "0.0.1"
    valid_formats = ["xlsx"]
    sheet_name_cols = []

    def _workbook_path(self, yaml_filename: str, wb_name: str = None):
        """Internal method that computes the path where the Excel workbook
        should be stored.

        :param yaml_filename: Name of provided LinkML schema
        :type yaml_filename: str
        :param wb_name: Prefix for the generated Excel spreadsheet name
        :type wb_name: str
        """
        # handle the case when an output filename is not provided
        if not wb_name:
            prefix, _ = os.path.splitext(os.path.basename(yaml_filename))
            prefix_root, prefix_ext = os.path.splitext(prefix)

            if prefix_ext == ".yaml":
                prefix = prefix_root

            output_xlsx = (
                f"{prefix}_{self.generator_name}_{self.generator_version}.xlsx"
            )

            return output_xlsx

        return wb_name

    def __init__(
        self,
        schema: Union[str, TextIO, SchemaDefinition],
        output: Optional[str] = None,
        **kwargs,
    ) -> None:
        self.wb_name = self._workbook_path(yaml_filename=schema, wb_name=output)
        self.workbook = Workbook()
        self.workbook.remove(self.workbook["Sheet"])

        # dictionary with slot types and possibles values for those types
        self.enum_dict = {}
        super().__init__(schema, **kwargs)

    def _create_spreadsheet(self, ws_name: str, columns: List[str]) -> None:
        """Method to add worksheets to the Excel workbook.

        :param ws_name: Name of each of the worksheets
        :type ws_name: str
        :param columns: Columns that are relevant to each of the worksheets
        :type columns: List[str]
        """
        ws = self.workbook.create_sheet(ws_name)
        self.workbook.active = ws
        ws.append(columns)
        self.workbook.save(self.wb_name)

    def visit_class(self, cls: ClassDefinition) -> bool:
        """Overridden method to intercept classes from generator framework."""
        self._create_spreadsheet(ws_name=camelcase(cls.name), columns=cls.slots)

        return True

    def visit_enum(self, enum: EnumDefinition) -> bool:
        """Overridden method to intercept enums from generator framework."""

        def extract_permissible_text(pv):
            if type(pv) is str:
                return pv
            if type(pv) is PermissibleValue:
                return pv.text.code
            if type(pv) is PermissibleValueText:
                return pv
            raise ValueError(f"Invalid permissible value in enum {enum}: {pv}")

        permissible_values_texts = list(
            map(extract_permissible_text, enum.permissible_values or [])
        )

        self.enum_dict[enum.name] = permissible_values_texts

    def visit_class_slot(
        self, cls: ClassDefinition, aliased_slot_name: str, slot: SlotDefinition
    ) -> None:
        """Overridden method to intercept classes and associated slots from generator 
        framework."""
        wb = load_workbook(self.wb_name)

        if cls.name in self.workbook.sheetnames:
            if slot.range in self.enum_dict:

                # TODO: the below list needs to be dynamically populated with items 
                # in dropdown
                # basically the values associated with an enum type
                valid = '"The,earth,revolves,around,sun"'

                ws = self.workbook[cls.name]

                dv = DataValidation(type="list", formula1=valid, allow_blank=True)
                ws.add_data_validation(dv)

                # TODO: select the columns from the sheet to which the dropdown is 
                # to be applied
                dv.add("B1")

                wb.save(self.wb_name)


@shared_arguments(ExcelGenerator)
@click.command()
@click.option(
    "-o",
    "--output",
    type=click.Path(),
    help="""Name of Excel spreadsheet to be created""",
)
def cli(yamlfile, **kwargs):
    """Generate Excel representation of a LinkML model"""
    print(ExcelGenerator(yamlfile, **kwargs).serialize(**kwargs))


if __name__ == "__main__":
    cli()
